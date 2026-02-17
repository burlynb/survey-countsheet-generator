"""Generate sea otter survey count sheet template from field log summaries."""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re
import sys

YEAR = int(input("Enter survey year: "))
INPUT_DIR = os.path.join(os.path.dirname(__file__), "inputs")
SITES_FILE = os.path.join(INPUT_DIR, "SITES.xlsx")
LOG_FILE = os.path.join(INPUT_DIR, f"{YEAR}_LOGSummary.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"COUNTSHEET_TEMPLATE_{YEAR}.xlsx")

# --- 1. Input Validation ---
for f, label in [(SITES_FILE, "SITES.xlsx"), (LOG_FILE, f"{YEAR}_LOGSummary.xlsx")]:
    if not os.path.exists(f):
        print(f"Error: Required file not found: {label}")
        sys.exit(1)

# --- 2. Data Loading ---
sites = pd.read_excel(SITES_FILE)
log = pd.read_excel(LOG_FILE)

print(f"SITES columns: {list(sites.columns)}")
print(f"LOGSummary columns (raw): {list(log.columns)}")
print(f"SITES rows: {len(sites)}")
print(f"LOGSummary rows: {len(log)}")

# Normalize LOGSummary column names to handle year-to-year variations
col_map = {
    "Date": "DATE",
    "Time": "TIME",
    "Count": "COUNT",
    "Pass": "PASS",
    "Start Frame": "START FRAME",
    "End Frame": "END FRAME",
    "Pass Description": "PASS DESCRIPTION",
    "Add": "ADD",
    "Disturbance?": "DISTURBANCE",
    "Disturbance": "DISTURBANCE",
    "PRIOR": "Priority",
    "Priority": "Priority",
    "PRIORITY": "Priority",
    "REG": "REGION",
    "REGION": "REGION",
    "REGNO": "REGNO",
}
log.rename(columns={k: v for k, v in col_map.items() if k in log.columns}, inplace=True)
print(f"LOGSummary columns (normalized): {list(log.columns)}")

# Normalize SUBSITE for matching
sites["_subsite_key"] = sites["SUBSITE"].astype(str).str.strip().str.upper()
log["_subsite_key"] = log["SUBSITE"].astype(str).str.strip().str.upper()

# --- 3A. Handle Duplicate Surveys (Multiple Passes) ---
# Remove "DO NOT USE" rows
log_clean = log[~log["SUBSITE"].astype(str).str.contains("DO NOT USE", case=False, na=False)].copy()

log_clean["DATE"] = pd.to_datetime(log_clean["DATE"], errors="coerce")
# Convert TIME to numeric for sorting
log_clean["_time_num"] = pd.to_numeric(log_clean["TIME"], errors="coerce")

def _concat_add(series):
    """Concatenate ADD values with ' + ', converting to integers."""
    vals = []
    for v in series:
        if pd.notna(v) and str(v).strip():
            # Strip non-numeric prefix like "Add " and convert to int
            s = str(v).strip()
            num = re.sub(r"[^\d]", "", s)
            if num:
                vals.append(str(int(num)))
            else:
                vals.append(s)
    return " + ".join(vals) if vals else ""

def _concat_disturbance(series):
    """Concatenate disturbance values with ' + ', removing repeated 'Disturbed' prefix."""
    vals = [str(v).strip() for v in series if pd.notna(v) and str(v).strip()]
    if not vals:
        return ""
    parts = []
    for v in vals:
        # Strip "Disturbed " prefix for all but the first entry
        if parts and v.lower().startswith("disturbed "):
            parts.append(v[len("Disturbed "):])
        else:
            parts.append(v)
    return " + ".join(parts)

def _concat_non_null(series):
    """Concatenate non-null values with '; ' separator."""
    vals = [str(v).strip() for v in series if pd.notna(v) and str(v).strip()]
    return "; ".join(vals) if vals else ""

# For each subsite group: aggregate multiple passes into one row
aggregated_rows = []
for key, group in log_clean.groupby("_subsite_key"):
    group_sorted = group.sort_values("_time_num", ascending=True)

    if len(group_sorted) == 1:
        base = group_sorted.iloc[0].copy()
    else:
        # Determine if there's a mix of photo passes (PASS not null) and count passes (COUNT not null)
        photo_rows = group_sorted[group_sorted["PASS"].notna() & (group_sorted["PASS"].astype(str).str.strip() != "")]
        count_rows = group_sorted[group_sorted["COUNT"].notna()]
        has_photo = len(photo_rows) > 0
        has_count = len(count_rows) > 0

        if has_photo and has_count:
            # Mixed photo/count: use the photo pass as base
            base = photo_rows.iloc[0].copy()
            # Add count values from count passes into ADD as "COUNT x"
            count_adds = []
            for _, cr in count_rows.iterrows():
                c = cr["COUNT"]
                count_adds.append(f"COUNT {int(c)}")
            existing_add = _concat_add(group_sorted["ADD"])
            all_adds = " + ".join([a for a in [existing_add] + count_adds if a])
            base["ADD"] = all_adds
        elif has_count and not has_photo:
            # Multiple count passes: use earliest time as base, concatenate counts
            base = group_sorted.iloc[0].copy()
            counts = [str(int(c)) for c in count_rows["COUNT"] if pd.notna(c)]
            base["COUNT"] = "+".join(counts) if counts else ""
        else:
            # Multiple photo passes: use earliest time as base
            base = group_sorted.iloc[0].copy()

    # Always aggregate ADD, DISTURBANCE, PASS DESCRIPTION across all passes
    if len(group_sorted) > 1:
        if not (has_photo and has_count):
            # Only aggregate ADD if we didn't already handle it in mixed mode
            base["ADD"] = _concat_add(group_sorted["ADD"])
        base["DISTURBANCE"] = _concat_disturbance(group_sorted["DISTURBANCE"])
        base["PASS DESCRIPTION"] = _concat_non_null(group_sorted["PASS DESCRIPTION"])

    # Combine MML_IDs if they differ (e.g., "183A" and "183B" -> "183A-B")
    if len(group_sorted) > 1:
        mml_ids = [str(m).strip() for m in group_sorted["MML_ID"] if pd.notna(m) and str(m).strip()]
        if len(set(mml_ids)) > 1:
            # Extract common numeric prefix and combine suffixes
            first = mml_ids[0]
            prefix_match = re.match(r"(\d+)", first)
            if prefix_match:
                prefix = prefix_match.group(1)
                suffixes = []
                for m in mml_ids:
                    suffix = m[len(prefix):]
                    if suffix:
                        suffixes.append(suffix)
                if suffixes:
                    base["MML_ID"] = prefix + "-".join(suffixes)
            else:
                base["MML_ID"] = mml_ids[0]

    aggregated_rows.append(base)

log_clean = pd.DataFrame(aggregated_rows)

# Track MML_IDs that were combined into another line (e.g., 183B is part of 183A-B)
# These individual MML_IDs are accounted for in a surveyed line already
combined_mml_ids = set()
for _, row in log_clean.iterrows():
    mml = str(row.get("MML_ID", "")).strip()
    if "-" in mml:
        # e.g., "183A-B" -> extract all individual IDs: "183A", "183B"
        prefix_match = re.match(r"(\d+)", mml)
        if prefix_match:
            prefix = prefix_match.group(1)
            rest = mml[len(prefix):]  # e.g., "A-B"
            suffixes = rest.split("-")
            for s in suffixes:
                combined_mml_ids.add(prefix + s)

# Build set of log subsites
log_subsites = set(log_clean["_subsite_key"])
sites_subsites = set(sites["_subsite_key"])

# --- 3B-D: Build output rows ---
rows = []

# Process all SITES entries
for _, site_row in sites.iterrows():
    key = site_row["_subsite_key"]
    log_match = log_clean[log_clean["_subsite_key"] == key]
    has_log = len(log_match) > 0
    log_row = log_match.iloc[0] if has_log else None

    # Survey status
    if has_log and pd.notna(log_row["DATE"]):
        survey = "OTTER"
    elif has_log and pd.isna(log_row["DATE"]):
        survey = "MISSED"
    else:
        # Check if this site's MML_ID is accounted for in a combined MML_ID line
        site_mml = str(site_row.get("MML_ID", "")).strip()
        if site_mml and site_mml in combined_mml_ids:
            survey = "SUBSITE"
        else:
            survey = "OUTSIDE"

    # COUNTTYPE
    counttype = ""
    if survey == "OTTER":
        if has_log and pd.notna(log_row.get("COUNT")) and str(log_row.get("COUNT", "")).strip() != "":
            counttype = 4
        elif has_log and pd.notna(log_row.get("PASS")) and str(log_row.get("PASS", "")).strip() != "":
            counttype = 3

    # PHOTO
    photo = ""
    if survey == "OTTER":
        if has_log and pd.notna(log_row.get("PASS")) and str(log_row.get("PASS", "")).strip() != "":
            photo = "Y"
        else:
            photo = "N"

    # FLAGS
    flags = ""
    flag_reason = ""
    if has_log:
        site_mml = str(site_row.get("MML_ID", "")).strip()
        log_mml = str(log_row.get("MML_ID", "")).strip()
        # MML_ID = "NEW" means new site
        if log_mml.upper() == "NEW":
            flags = "NEW SITE"
            flag_reason = "MML_ID marked as NEW"
        else:
            # Compare only numeric prefix (e.g., "248A" -> "248")
            site_mml_num = re.match(r"(\d+)", site_mml)
            log_mml_num = re.match(r"(\d+)", log_mml)
            site_mml_num = site_mml_num.group(1) if site_mml_num else site_mml
            log_mml_num = log_mml_num.group(1) if log_mml_num else log_mml
            if site_mml_num and log_mml_num and site_mml_num != log_mml_num:
                flags = "NEEDS_REVIEW"
                flag_reason = f"MML_ID mismatch: SITES={site_mml}, LOG={log_mml}"

    def _val(series_row, col):
        """Get value from row, return blank string if null."""
        v = series_row.get(col)
        if pd.isna(v):
            return ""
        return v

    row = {
        "FLAGS": flags,
        "_flag_reason": flag_reason,
        "_sites_mml": str(site_row.get("MML_ID", "")).strip(),
        "_log_mml": str(log_row.get("MML_ID", "")).strip() if has_log else "",
        "SUBSITE": log_row["SUBSITE"] if has_log else site_row["SUBSITE"],
        "SUBSITE_ID": _val(site_row, "SUBSITE_ID"),
        "PARENTSITE": _val(site_row, "PARENTSITE"),
        "PARENTSITE_ID": _val(site_row, "PARENTSITE_ID"),
        "MML_ID": _val(log_row, "MML_ID") if (has_log and survey == "OTTER") else _val(site_row, "MML_ID"),
        "REGION": _val(log_row, "REGION") if (has_log and survey in ("OTTER", "MISSED")) else _val(site_row, "REGION"),
        "REGNO": _val(log_row, "REGNO") if (has_log and survey in ("OTTER", "MISSED")) else _val(site_row, "REGNO"),
        "RCA": _val(log_row, "RCA") if (has_log and survey in ("OTTER", "MISSED")) else _val(site_row, "RCA"),
        "ROOK": _val(log_row, "ROOK") if (has_log and survey in ("OTTER", "MISSED")) else _val(site_row, "ROOK"),
        "LAT": _val(site_row, "LAT"),
        "LON": _val(site_row, "LON"),
        "PRIORITY": _val(log_row, "Priority") if has_log else "",
        "DATE": _val(log_row, "DATE") if has_log else "",
        "SURVEY": survey,
        "COUNTTYPE": counttype,
        "TIME": _val(log_row, "TIME") if has_log else "",
        "PHOTO": photo,
        "LOG_COUNT": _val(log_row, "COUNT") if has_log else "",
        "ADD": _val(log_row, "ADD") if has_log else "",
        "FRAME": "",
        "BULL": "",
        "SAM": "",
        "FEM": "",
        "JUV": "",
        "PUP": "",
        "PUP_DEAD": "",
        "NP_DEAD": "",
        "NP_TOTAL": "",
        "ALL_COUNT": "",
        "COUNTER_NOTES": "",
        "DISTURBANCE": _val(log_row, "DISTURBANCE") if has_log else "",
        "BRANDS": "",
        "COUNTER": "",
        "SURVEY NOTES": _val(log_row, "PASS DESCRIPTION") if has_log else "",
    }
    rows.append(row)

# Check for NEW SITE entries (in log but not in sites)
new_site_keys = log_subsites - sites_subsites
# Remove any NaN/empty keys
new_site_keys = {k for k in new_site_keys if isinstance(k, str) and k.strip() and k != "NAN"}
print(f"New site keys ({len(new_site_keys)}): {new_site_keys}")
for key in new_site_keys:
    log_row = log_clean[log_clean["_subsite_key"] == key].iloc[0]
    survey = "OTTER" if pd.notna(log_row["DATE"]) else "MISSED"

    counttype = ""
    if survey == "OTTER":
        if pd.notna(log_row.get("COUNT")) and str(log_row.get("COUNT", "")).strip() != "":
            counttype = 4
        elif pd.notna(log_row.get("PASS")) and str(log_row.get("PASS", "")).strip() != "":
            counttype = 3

    photo = ""
    if survey == "OTTER":
        if pd.notna(log_row.get("PASS")) and str(log_row.get("PASS", "")).strip() != "":
            photo = "Y"
        else:
            photo = "N"

    row = {
        "FLAGS": "NEW SITE",
        "_flag_reason": "SUBSITE not in SITES.xlsx",
        "_sites_mml": "",
        "_log_mml": str(_val(log_row, "MML_ID")),
        "SUBSITE": log_row["SUBSITE"],
        "SUBSITE_ID": "",
        "PARENTSITE": "",
        "PARENTSITE_ID": "",
        "MML_ID": _val(log_row, "MML_ID"),
        "REGION": _val(log_row, "REGION"),
        "REGNO": _val(log_row, "REGNO"),
        "RCA": _val(log_row, "RCA"),
        "ROOK": _val(log_row, "ROOK"),
        "LAT": "",
        "LON": "",
        "PRIORITY": _val(log_row, "Priority"),
        "DATE": _val(log_row, "DATE"),
        "SURVEY": survey,
        "COUNTTYPE": counttype,
        "TIME": _val(log_row, "TIME"),
        "PHOTO": photo,
        "LOG_COUNT": _val(log_row, "COUNT"),
        "ADD": _val(log_row, "ADD"),
        "FRAME": "",
        "BULL": "",
        "SAM": "",
        "FEM": "",
        "JUV": "",
        "PUP": "",
        "PUP_DEAD": "",
        "NP_DEAD": "",
        "NP_TOTAL": "",
        "ALL_COUNT": "",
        "COUNTER_NOTES": "",
        "DISTURBANCE": _val(log_row, "DISTURBANCE"),
        "BRANDS": "",
        "COUNTER": "",
        "SURVEY NOTES": _val(log_row, "PASS DESCRIPTION"),
    }
    rows.append(row)

# --- 5. Build DataFrame ---
df = pd.DataFrame(rows)

# Sort by SURVEY (OTTER → MISSED → OUTSIDE), then DATE (earliest → latest), then SUBSITE (A→Z)
survey_order = {"OTTER": 0, "MISSED": 1, "SUBSITE": 2, "OUTSIDE": 3}
df["_sort_survey"] = df["SURVEY"].map(survey_order)
df["_sort_date"] = pd.to_datetime(df["DATE"], errors="coerce")
df["_sort_subsite"] = df["SUBSITE"].astype(str).str.upper()
df = df.sort_values(["_sort_survey", "_sort_date", "_sort_subsite"], na_position="last").drop(columns=["_sort_survey", "_sort_date", "_sort_subsite"])

# --- 6. Output Generation ---

# Generate error report for flagged sites
flagged = df[df["FLAGS"] != ""].copy()
if len(flagged) > 0:
    error_report = flagged[["FLAGS", "SUBSITE", "_sites_mml", "_log_mml", "_flag_reason"]].copy()
    error_report.columns = ["FLAGS", "SUBSITE", "SITES_MML_ID", "LOG_MML_ID", "REASON"]
    error_report_file = os.path.join(OUTPUT_DIR, f"FLAGGED_SITES_{YEAR}.xlsx")
    error_report.to_excel(error_report_file, index=False, engine="openpyxl")
    # Format flagged sites report
    wb_err = load_workbook(error_report_file)
    ws_err = wb_err.active
    center_align = Alignment(horizontal="center")
    for cell in ws_err[1]:
        cell.font = Font(bold=True)
    # Center SITES_MML_ID and LOG_MML_ID columns (columns C and D)
    for col_idx in [3, 4]:
        for row in ws_err.iter_rows(min_row=1, max_row=ws_err.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = center_align
    # Auto-fit columns
    for col_idx, col_cells in enumerate(ws_err.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws_err.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    wb_err.save(error_report_file)
    print(f"Error report saved: {error_report_file}")

# Drop internal columns before writing main output
df = df.drop(columns=["_flag_reason", "_sites_mml", "_log_mml"])
df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

# Apply formatting
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# Bold header
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Freeze top row
ws.freeze_panes = "A2"

# Auto-fit column widths
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_len = 0
    for cell in col_cells:
        val = str(cell.value) if cell.value else ""
        max_len = max(max_len, len(val))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

# Highlight flagged rows in yellow
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
flags_col_idx = 1  # FLAGS is column A
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value and str(row[0].value).strip():
        for cell in row:
            cell.fill = yellow_fill

# Center MML_ID column
center_align = Alignment(horizontal="center")
mml_col_idx = list(df.columns).index("MML_ID") + 1  # 1-based
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=mml_col_idx, max_col=mml_col_idx):
    for cell in row:
        cell.alignment = center_align

# Format DATE column as m/dd display
date_col_idx = list(df.columns).index("DATE") + 1  # 1-based
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=date_col_idx, max_col=date_col_idx):
    for cell in row:
        if cell.value:
            cell.number_format = "m/dd"

wb.save(OUTPUT_FILE)

# --- 7. Summary Report ---
total = len(df)
otter_count = len(df[df["SURVEY"] == "OTTER"])
missed_count = len(df[df["SURVEY"] == "MISSED"])
subsite_count = len(df[df["SURVEY"] == "SUBSITE"])
outside_count = len(df[df["SURVEY"] == "OUTSIDE"])
new_site_count = len(df[df["FLAGS"] == "NEW SITE"])
review_count = len(df[df["FLAGS"] == "NEEDS_REVIEW"])

print(f"""
Count Sheet Generation Summary for {YEAR}
==========================================
Total sites in template: {total}
  - OTTER (surveyed): {otter_count}
  - MISSED (planned but not surveyed): {missed_count}
  - SUBSITE (included in another surveyed line): {subsite_count}
  - OUTSIDE (not planned): {outside_count}

Flags raised:
  - NEW SITE: {new_site_count} sites
  - NEEDS_REVIEW: {review_count} sites

Column count: {len(df.columns)} (expected 35)

Output file: {OUTPUT_FILE}
""")

# Quality checks
assert df["SURVEY"].isin(["OTTER", "MISSED", "OUTSIDE", "SUBSITE"]).all(), "Invalid SURVEY values found"
assert df["COUNTTYPE"].isin(["", 3, 4]).all(), "Invalid COUNTTYPE values found"
assert len(df.columns) == 35, f"Expected 35 columns, got {len(df.columns)}"

# Check no duplicate subsites
dup_check = df[df["SUBSITE"].astype(str).str.strip().str.upper().duplicated(keep=False)]
if len(dup_check) > 0:
    print(f"WARNING: {len(dup_check)} duplicate SUBSITE entries found")
    print(dup_check[["FLAGS", "SUBSITE", "SURVEY"]].to_string())

print("Count sheet generation complete.")
