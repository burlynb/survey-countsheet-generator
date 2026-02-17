#!/usr/bin/env python3
"""
Sea Otter Survey Count Sheet Generator
Implements the generate-countsheet skill logic
"""

import pandas as pd
import sys
from pathlib import Path
from datetime import datetime


def load_sites(filepath='SITES.xlsx'):
    """Load master sites list"""
    try:
        sites = pd.read_excel(filepath)
        required_cols = ['SUBSITE', 'SUBSITE_ID', 'PARENTSITE', 'PARENTSITE_ID', 
                        'MML_ID', 'REGION', 'REGNO', 'RCA', 'ROOK', 'LAT', 'LON']
        
        missing = [col for col in required_cols if col not in sites.columns]
        if missing:
            raise ValueError(f"Missing columns in SITES.xlsx: {missing}")
        
        # Clean SUBSITE values
        sites['SUBSITE'] = sites['SUBSITE'].str.strip()
        
        print(f"✓ Loaded {len(sites)} sites from SITES.xlsx")
        return sites
    except FileNotFoundError:
        print(f"✗ Error: SITES.xlsx not found in current directory")
        sys.exit(1)
    except Exception as e:
        print(f"✗ Error loading SITES.xlsx: {e}")
        sys.exit(1)


def load_logsummary(year):
    """Load year-specific log summary"""
    filepath = f"{year}_LOGSummary.xlsx"
    try:
        log = pd.read_excel(filepath)
        required_cols = ['DATE', 'MML_ID', 'SUBSITE', 'PARENTSITE', 'TIME', 
                        'COUNT', 'PASS', 'PASS DESCRIPTION', 'ADD', 'DISTURBANCE',
                        'Priority', 'REGION', 'REGNO', 'RCA', 'ROOK']
        
        missing = [col for col in required_cols if col not in log.columns]
        if missing:
            raise ValueError(f"Missing columns in {filepath}: {missing}")
        
        # Clean SUBSITE values
        log['SUBSITE'] = log['SUBSITE'].str.strip()
        
        print(f"✓ Loaded {len(log)} entries from {filepath}")
        return log
    except FileNotFoundError:
        print(f"✗ Error: {filepath} not found in current directory")
        print(f"   Expected format: YYYY_LOGSummary.xlsx (e.g., 2024_LOGSummary.xlsx)")
        sys.exit(1)
    except Exception as e:
        print(f"✗ Error loading {filepath}: {e}")
        sys.exit(1)


def remove_do_not_use(log):
    """Remove entries marked as 'DO NOT USE'"""
    before = len(log)
    log = log[~log['SUBSITE'].str.contains('DO NOT USE', case=False, na=False)]
    removed = before - len(log)
    if removed > 0:
        print(f"✓ Removed {removed} 'DO NOT USE' entries")
    return log


def handle_duplicates(log):
    """Handle duplicate surveys of the same site"""
    duplicates = log[log.duplicated(subset=['SUBSITE'], keep=False)]
    
    if len(duplicates) > 0:
        print(f"⚠ Found {len(duplicates)} duplicate entries for {duplicates['SUBSITE'].nunique()} sites")
        # Keep most recent date for each SUBSITE
        log = log.sort_values('DATE', ascending=False)
        log = log.drop_duplicates(subset=['SUBSITE'], keep='first')
        print(f"✓ Kept most recent entry for each duplicate site")
    
    return log


def determine_survey_status(row, sites_subsites):
    """Determine SURVEY status for a row"""
    subsite = row['SUBSITE']
    has_date = pd.notna(row.get('DATE'))
    
    if pd.notna(has_date) and has_date:
        return 'OTTER'  # Successfully surveyed
    elif subsite in sites_subsites:
        return 'MISSED'  # Planned but not surveyed
    else:
        return 'OUTSIDE'  # Not in SITES (shouldn't happen after merge)


def calculate_counttype(row):
    """Calculate COUNTTYPE based on COUNT and PASS columns"""
    if row['SURVEY'] == 'OTTER':
        if pd.notna(row.get('LOG_COUNT')):
            return 4  # Visual count
        elif pd.notna(row.get('PASS')):
            return 3  # Photographic count
    return None


def calculate_photo(row):
    """Calculate PHOTO value"""
    if pd.notna(row.get('PASS')):
        return 'Y'
    return None


def generate_flags(row, sites_dict):
    """Generate FLAGS for quality checks"""
    flags = []
    subsite = row['SUBSITE']
    
    # Check for NEW SITE
    if subsite not in sites_dict and row['SURVEY'] in ['OTTER', 'MISSED']:
        flags.append('NEW SITE')
    
    # Check for MML_ID mismatch
    if subsite in sites_dict:
        sites_mml = sites_dict[subsite].get('MML_ID')
        log_mml = row.get('MML_ID_log')
        if pd.notna(log_mml) and pd.notna(sites_mml):
            if str(sites_mml) != str(log_mml):
                flags.append('NEEDS_REVIEW')
    
    return ', '.join(flags) if flags else None


def merge_data(sites, log):
    """Merge SITES and LOGSUMMARY data according to rules"""
    print("\n=== Merging Data ===")
    
    # Create lookup dictionaries
    sites_dict = sites.set_index('SUBSITE').to_dict('index')
    sites_subsites = set(sites['SUBSITE'].unique())
    
    # Start with all sites from SITES
    result = sites.copy()
    result['FROM_LOG'] = False
    
    # Add log data
    log_copy = log.copy()
    log_copy['FROM_LOG'] = True
    
    # Merge on SUBSITE
    # First, prepare log columns with _log suffix for comparison
    log_merge = log_copy.copy()
    log_merge = log_merge.rename(columns={'MML_ID': 'MML_ID_log'})
    
    # Merge
    merged = result.merge(
        log_merge[['SUBSITE', 'DATE', 'TIME', 'COUNT', 'PASS', 'PASS DESCRIPTION',
                  'ADD', 'DISTURBANCE', 'Priority', 'MML_ID_log', 'FROM_LOG',
                  'REGION', 'REGNO', 'RCA', 'ROOK']],
        on='SUBSITE',
        how='outer',
        suffixes=('', '_log')
    )
    
    # Determine SURVEY status
    merged['SURVEY'] = merged.apply(
        lambda row: determine_survey_status(row, sites_subsites), axis=1
    )
    
    # For sites that were surveyed, use log data for REGION, REGNO, RCA, ROOK
    for col in ['REGION', 'REGNO', 'RCA', 'ROOK']:
        log_col = f'{col}_log'
        if log_col in merged.columns:
            merged[col] = merged.apply(
                lambda row: row[log_col] if pd.notna(row[log_col]) else row[col],
                axis=1
            )
    
    # Rename columns for output
    merged = merged.rename(columns={
        'COUNT': 'LOG_COUNT',
        'PASS DESCRIPTION': 'SURVEY NOTES'
    })
    
    # Calculate COUNTTYPE and PHOTO
    merged['COUNTTYPE'] = merged.apply(calculate_counttype, axis=1)
    merged['PHOTO'] = merged.apply(calculate_photo, axis=1)
    
    # Generate FLAGS
    merged['FLAGS'] = merged.apply(lambda row: generate_flags(row, sites_dict), axis=1)
    
    # Add blank columns for manual entry
    manual_cols = ['FRAME', 'BULL', 'SAM', 'FEM', 'JUV', 'PUP', 'PUP_DEAD',
                   'NP_DEAD', 'NP_TOTAL', 'ALL_COUNT', 'COUNTER_NOTES', 
                   'BRANDS', 'COUNTER']
    for col in manual_cols:
        merged[col] = None
    
    # Reorder columns according to specification
    output_cols = [
        'FLAGS', 'SUBSITE', 'SUBSITE_ID', 'PARENTSITE', 'PARENTSITE_ID', 'MML_ID',
        'REGION', 'REGNO', 'RCA', 'ROOK', 'LAT', 'LON', 'Priority', 'DATE', 'SURVEY',
        'COUNTTYPE', 'TIME', 'PHOTO', 'LOG_COUNT', 'ADD', 'FRAME', 'BULL', 'SAM',
        'FEM', 'JUV', 'PUP', 'PUP_DEAD', 'NP_DEAD', 'NP_TOTAL', 'ALL_COUNT',
        'COUNTER_NOTES', 'DISTURBANCE', 'BRANDS', 'COUNTER', 'SURVEY NOTES'
    ]
    
    # Keep only columns that exist
    output_cols = [col for col in output_cols if col in merged.columns]
    result_df = merged[output_cols]
    
    # Sort by REGION, then SUBSITE
    result_df = result_df.sort_values(['REGION', 'SUBSITE'])
    
    return result_df


def generate_summary(df, year):
    """Generate summary statistics"""
    print("\n" + "="*50)
    print(f"Count Sheet Generation Summary for {year}")
    print("="*50)
    print(f"Total sites in template: {len(df)}")
    
    survey_counts = df['SURVEY'].value_counts()
    for status in ['OTTER', 'MISSED', 'OUTSIDE']:
        count = survey_counts.get(status, 0)
        print(f"  - {status}: {count}")
    
    print(f"\nFlags raised:")
    flag_counts = df['FLAGS'].notna().sum()
    if flag_counts > 0:
        new_site = df['FLAGS'].str.contains('NEW SITE', na=False).sum()
        needs_review = df['FLAGS'].str.contains('NEEDS_REVIEW', na=False).sum()
        print(f"  - NEW SITE: {new_site}")
        print(f"  - NEEDS_REVIEW: {needs_review}")
    else:
        print(f"  - No flags")
    
    return survey_counts


def main():
    """Main execution function"""
    # Determine year from command line or prompt
    if len(sys.argv) > 1:
        year = sys.argv[1]
    else:
        year = input("Enter survey year (e.g., 2024): ").strip()
    
    if not year.isdigit():
        print("Error: Year must be numeric (e.g., 2024)")
        sys.exit(1)
    
    print(f"\n{'='*50}")
    print(f"Sea Otter Count Sheet Generator - {year}")
    print(f"{'='*50}\n")
    
    # Load data
    sites = load_sites()
    log = load_logsummary(year)
    
    # Process data
    log = remove_do_not_use(log)
    log = handle_duplicates(log)
    
    # Merge and transform
    result = merge_data(sites, log)
    
    # Generate output
    output_file = f"COUNTSHEET_TEMPLATE_{year}.xlsx"
    result.to_excel(output_file, index=False, engine='openpyxl')
    
    # Format Excel file
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        # Highlight flagged rows in yellow
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value:  # FLAGS column has content
                for cell in row:
                    cell.fill = yellow_fill
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        print(f"\n✓ Applied Excel formatting")
    except Exception as e:
        print(f"\n⚠ Warning: Could not apply Excel formatting: {e}")
    
    # Generate summary
    summary = generate_summary(result, year)
    
    print(f"\n{'='*50}")
    print(f"✓ Output file: {output_file}")
    print(f"{'='*50}\n")


if __name__ == "__main__":
    main()
